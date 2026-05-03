import os
import numpy as np
import jax.numpy as jnp
from openpyxl import Workbook, load_workbook
import json
from pathlib import Path
# import rospy
# import rosbag
import yaml
from geometry_msgs.msg import PoseStamped, TwistStamped

PACKAGE_ROOT = Path(__file__).resolve().parents[1]
with open(PACKAGE_ROOT / "datas" / "config" / "config.yaml", "r") as f:
    loaded = yaml.safe_load(f)
# 提取 opt_args 子字典
config = loaded["opt_args"]
dt = config["dt"]
def export_map_to_jsonl(map_info, log_file=None):
    mapinfo_to_save = {
        "means": [m.tolist() for m in map_info['means']],
        "covs": [c.tolist() for c in map_info['covs']],
    }
    with open(log_file, "a") as f:
        f.write(json.dumps(mapinfo_to_save) + "\n")

def load_map_history_jsonl(file_path="map_history.jsonl"):
    history = []
    with open(file_path, "r") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            record = json.loads(line)
            # 转换 means 和 covs 为 jnp.array 列表
            converted_record = {
                "means": [jnp.array(m) for m in record["means"]],
                "covs": [jnp.array(c) for c in record["covs"]],
            }
            history.append(converted_record)
    return history
def save_ergodic_metrics_to_excel(robot_num, global_metric, decay_type, file_path='ergodic_metric.xlsx'):
    """
    将每个机器人的 ergodic metric 数据及其 type 汇总写入 Excel 文件。
    
    参数:
        global_metric (dict): 键为 robot_id，值为包含 'values' 列表的字典。
        decay_type (str): 当前实验类型（如 'noexchange'）。
        file_path (str): Excel 文件路径，默认为 'ergodic_metric.xlsx'。
    """
    # 确保文件存在并包含正确的工作表和表头
    if not os.path.exists(file_path):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "ergodic_metric"
        header = ["type", "metric_mean", "mse"]
        sheet.append(header)
        wb.save(file_path)

    # 加载工作簿
    wb = load_workbook(file_path)
    if "ergodic_metric" in wb.sheetnames:
        sheet = wb["ergodic_metric"]
    else:
        sheet = wb.active
        sheet.title = "ergodic_metric"
        header = ["type", "metric_mean", "mse"]
        sheet.append(header)

    # 准备数据
    metric_means = []
    mses = []
    new_row_data = [None, None, None, None]
    values_list = global_metric['values']
    mean_value = float(np.mean(values_list))
    variance_value = float(np.var(values_list))  # 若需 MSE，请替换为相应计算

    new_row_data[0] = str(decay_type)
    new_row_data[1] = str(mean_value)
    new_row_data[2] = str(variance_value)
    sheet.append(new_row_data.copy())
    # 保存文件
    wb.save(file_path)
    print(f"数据已成功追加到 {file_path}")

def append_metric(global_metric, metric):
    global_metric['time'].append(len(global_metric['time']))
    global_metric['values'].append(metric)
            
def save_to_rosbag(traj, bagfile = "../bags/traj.bag"):
    with rosbag.Bag(bagfile, 'w') as bag:
        for i, point in enumerate(traj):
            pose = PoseStamped()
            pose.header.stamp = rospy.Time.now()
            pose.header.frame_id = "map"
            pose.pose.position.x = point.x
            pose.pose.position.y = point.y
            bag.write('/trajectory/pose', pose, pose.header.stamp)

            twist = TwistStamped()
            twist.header = pose.header
            twist.twist.linear.x = point.vx


def traj_to_rosbag(sol_trajs, saver, current_time):
    
    saver.save_traj2bag(
    trajectory_dict={'x':sol_trajs[0]['x'][:current_time], 'u':sol_trajs[0]['u'][:current_time]},
    bag_filename="traj0.bag",
    dt=dt,
    robot_id=0,
    )
    saver.save_traj2bag(
        trajectory_dict={'x':sol_trajs[1]['x'][:current_time], 'u':sol_trajs[1]['u'][:current_time]},
        bag_filename="traj1.bag",
        dt=dt,
        robot_id=1
    )
    saver.save_traj2bag(
        trajectory_dict={'x':sol_trajs[2]['x'][:current_time], 'u':sol_trajs[2]['u'][:current_time]},
        bag_filename="traj2.bag",
        dt=dt,
        robot_id=2
    )

    saver.save_traj2bag(
        trajectory_dict={'x':sol_trajs[3]['x'][:current_time], 'u':sol_trajs[3]['u'][:current_time]},
        bag_filename="traj3.bag",
        dt=dt,
        robot_id=3
    )

def multi_traj_to_rosbag(sol_trajs, saver, current_time):
    saver.save_traj2bag(
        trajectory_dict={'x':sol_trajs[0]['x'][:current_time, 0:4], 'u':sol_trajs[0]['u'][:current_time,0:2]},
        bag_filename=f"traj0.bag",
        dt=dt,
        robot_id=0,
        )
    saver.save_traj2bag(
        trajectory_dict={'x':sol_trajs[1]['x'][:current_time, 4:8], 'u':sol_trajs[1]['u'][:current_time, 2:4]},
        bag_filename=f"traj1.bag",
        dt=dt,
        robot_id=1,
        )
    saver.save_traj2bag(
        trajectory_dict={'x':sol_trajs[2]['x'][:current_time, 8:12], 'u':sol_trajs[2]['u'][:current_time, 4:6]},
        bag_filename=f"traj2.bag",
        dt=dt,
        robot_id=2,
        )
    saver.save_traj2bag(
        trajectory_dict={'x':sol_trajs[3]['x'][:current_time, 12:16], 'u':sol_trajs[3]['u'][:current_time, 6:8]},
        bag_filename=f"traj3.bag",
        dt=dt,
        robot_id=3,
        )
    for i in range(4):
        saver.save_traj2bag(
        trajectory_dict={'x':sol_trajs[i]['x'][:current_time, 0:4], 'u':sol_trajs[i]['u'][:current_time,0:2]},
        bag_filename=f"traj0-{i}.bag",
        dt=dt,
        robot_id=0,
        )
        saver.save_traj2bag(
            trajectory_dict={'x':sol_trajs[i]['x'][:current_time, 4:8], 'u':sol_trajs[i]['u'][:current_time, 2:4]},
            bag_filename=f"traj1-{i}.bag",
            dt=dt,
            robot_id=1,
        )
        saver.save_traj2bag(
            trajectory_dict={'x':sol_trajs[i]['x'][:current_time, 8:12], 'u':sol_trajs[i]['u'][:current_time, 4:6]},
            bag_filename=f"traj2-{i}.bag",
            dt=dt,
            robot_id=2,
        )

        saver.save_traj2bag(
            trajectory_dict={'x':sol_trajs[i]['x'][:current_time, 12:16], 'u':sol_trajs[i]['u'][:current_time, 6:8]},
            bag_filename=f"traj3-{i}.bag",
            dt=dt,
            robot_id=3,
        )

def path_to_rosbag(sol_trajs, Pathsaver, search_time):
    Pathsaver.save_path_to_bag(sol_trajs[0]['x'][search_time:,:2], bag_filename="path0.bag", dt=dt, robot_id=0,
        frame_id="world")
    Pathsaver.save_path_to_bag(sol_trajs[1]['x'][search_time:,:2], bag_filename="path1.bag", dt=dt, robot_id=1,
        frame_id="world")
    Pathsaver.save_path_to_bag(sol_trajs[2]['x'][search_time:,:2], bag_filename="path2.bag", dt=dt, robot_id=2,
        frame_id="world")
    Pathsaver.save_path_to_bag(sol_trajs[3]['x'][search_time:,:2], bag_filename="path3.bag", dt=dt, robot_id=3,
        frame_id="world")
def multi_path_to_rosbag(sol_trajs, Pathsaver, search_time):
    for i in range(search_time):
        Pathsaver.save_path_to_bag(sol_trajs[0]['x'][i:,:2], bag_filename=f"path0.bag", dt=dt, robot_id=0,
            frame_id="world")
        Pathsaver.save_path_to_bag(sol_trajs[1]['x'][i:,4:6], bag_filename=f"path1.bag", dt=dt, robot_id=1,
            frame_id="world")
        Pathsaver.save_path_to_bag(sol_trajs[2]['x'][i:,8:10], bag_filename=f"path2.bag", dt=dt, robot_id=2,
            frame_id="world")
        Pathsaver.save_path_to_bag(sol_trajs[3]['x'][i:,12:14], bag_filename=f"path3.bag", dt=dt, robot_id=3,
            frame_id="world")
        for j in range(4):
            Pathsaver.save_path_to_bag(sol_trajs[j]['x'][i:,:2], bag_filename=f"path0-{j}.bag", dt=dt, robot_id=0,
                frame_id="world")
            Pathsaver.save_path_to_bag(sol_trajs[j]['x'][i:,4:6], bag_filename=f"path1-{j}.bag", dt=dt, robot_id=1,
                frame_id="world")
            Pathsaver.save_path_to_bag(sol_trajs[j]['x'][i:,8:10], bag_filename=f"path2-{j}.bag", dt=dt, robot_id=2,
                frame_id="world")
            Pathsaver.save_path_to_bag(sol_trajs[j]['x'][i:,12:14], bag_filename=f"path3-{j}.bag", dt=dt, robot_id=3,
                frame_id="world")
        
def multi_map_to_rosbag(robot_distr, map_saver, update_map_freq):
    map_saver.save_probmap_to_bag(robot_distr[0].evals[1], robot_distr[0].evals[0], dt = dt, 
                                  timesteps=update_map_freq, bag_filename="map_r0.bag", color=["#ffffff", "#1D6CD4"])
    map_saver.save_probmap_to_bag(robot_distr[1].evals[1], robot_distr[1].evals[0], dt = dt, 
                                  timesteps=update_map_freq, bag_filename="map_r1.bag", color=["#ffffff", "#FFE100"])
    map_saver.save_probmap_to_bag(robot_distr[2].evals[1], robot_distr[2].evals[0], dt = dt, 
                                  timesteps=update_map_freq, bag_filename="map_r2.bag", color=["#ffffff", "#01A064"])
    map_saver.save_probmap_to_bag(robot_distr[3].evals[1], robot_distr[3].evals[0], dt = dt, 
                                  timesteps=update_map_freq, bag_filename="map_r3.bag", color=["#ffffff", "#AA00FF"])
