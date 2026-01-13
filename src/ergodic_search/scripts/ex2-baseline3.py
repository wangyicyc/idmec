
import os
os.environ["JAX_ENABLE_X64"] = "True"
import sys 
sys.path.append('..')
# 增广拉格朗日法优化器
from multiRobots_lib.solver import al_iLQR 
from multiRobots_lib.augument_lagrange_func import loss_traj_single, eq_constr_single, ineq_constr_single, loss_compare_single
from multiRobots_lib.data_collect import append_metric, traj_to_rosbag, path_to_rosbag
import logging
import numpy as np
from multiRobots_lib.plot_utils import plot_trajs_old
from multiRobots_lib.tools import exchange_info_old, optimize_trajs, update_accumulated_time
from multiRobots_lib.decay_utils import update_beta_single
from multiRobots_lib.data2bag import CommandToRosbag, MapINfoToMarkers, PathToRosbag
from IPython.display import clear_output
from multiRobots_lib.hyper_params import (
    opt_args,
    target_distr,
    robot_number,
    init_state_old,
    sol_trajs_old,
    robot_model_single,
    betas,
)

robot_distr = []
for robot_id in range(robot_number):
    robot_distr.append(target_distr)          # 目标分布对

logging.basicConfig(
    filename='datas/logs/app.log',          # 日志文件名
    level=logging.INFO,          # 日志等级
    format='%(asctime)s [%(levelname)s] %(message)s',  # 格式
)

# 创建保存器
# commandSaver = CommandToRosbag(bag_dir="../datas/bags/baseline1")   
# map_saver = MapINfoToMarkers(bag_dir="../datas/bags/baseline1", frame_id="world")
# pathSaver = PathToRosbag(bag_dir="../datas/bags/baseline1") 
# 保存完整轨迹

update_map_freq = opt_args["update_map_freq"]
map_merge_freq = opt_args["map_merge_freq"] 
init_state = init_state_old
sol_trajs = sol_trajs_old
robot_number = opt_args["robot_number"]
tsteps = opt_args["tsteps"]
# 双机器人初始化
start_pos = opt_args["x0"]
end_pos = opt_args["xf"]
# 假设你想创建包含 n 个 al_iLQR 结果的数组
n = robot_number  # 例如，创建 10 个结果
traj_solver = [al_iLQR(
    args=opt_args,
    objective=loss_traj_single,
    dynamics=robot_model_single,
    inequality=ineq_constr_single,
    equality=eq_constr_single,
    target_distr = robot_distr[_id].evals,
    robot_id = _id,
) for _id in range(n)]
decay_type = 'linear'
init_dual = True
save_path = '../figures/ex2-baseline3.png'
map_merge_cnt = 0
involved_robots = list(range(robot_number))
involved_robots = set(involved_robots)
accumulated_time = 0
be_num = 1
global_metric = {
        'time': [],
        'values': [],
}
current_time = None
robot_pair = []
logging.info('ex2-baseline3')

while True:
    sol_trajs, to_remove = optimize_trajs(involved_robots, sol_trajs, betas, traj_solver, init_state, init_dual)
    involved_robots -= to_remove
    clear_output(wait=True)                   # 清除上一次输出，动态刷新
    # init_dual = False
    # =================================== 更新地图 ==============================================================
    current_time, accumulated_time = update_accumulated_time(current_time, accumulated_time, be_num)
    map_merge_cnt += current_time
    if map_merge_cnt >= map_merge_freq:
        accumulated_time -= (map_merge_cnt - map_merge_freq)
        current_time -= (map_merge_cnt - map_merge_freq)
        map_merge_cnt = 0
        if accumulated_time >= tsteps:
            append_metric(global_metric, loss_compare_single(sol_trajs, target_distr.evals, current_time))
            break
        for r_id in range(robot_number):
            u_t = sol_trajs[r_id]['x'][current_time, 0:2]
    # ================================ 为replan准备 ============================================================== 
    if accumulated_time == update_map_freq * be_num:
        append_metric(global_metric, loss_compare_single(sol_trajs, target_distr.evals, current_time))
        if accumulated_time >= tsteps:
            break
        logging.info("update map")
        # map_saver.save_heatmap_to_bag(target_distr.evals[1], target_distr.evals[0], 0.12, update_map_freq)
        # target_distr.update_map(accumulated_time, "perturb", 'write')
        target_distr.update_map(accumulated_time, "perturb", 'read')
        be_num += 1

    sol_trajs, connected_pairs = exchange_info_old(sol_trajs, robot_pair, current_time, robot_distr)
    betas = update_beta_single(sol_trajs, decay_type)
    involved_robots = set(list(range(robot_number)))
    init_state = [traj['x'][0, :] for traj in sol_trajs]  
    init_dual = True
    logging.warning(f"connect time:{current_time}, accumulated time:{accumulated_time}")
    current_time = None
plot_trajs_old(start_pos, end_pos, sol_trajs, betas, robot_distr, save_path)  

from openpyxl import Workbook, load_workbook
def save_ergodic_metrics_to_excel(global_metric, decay_type, file_path='experiment2.xlsx'):
    values_list = global_metric['values']
    
    # 构造完整表头：type + 所有 metric 列（如 metric_0, metric_1, ...）
    max_cols_needed = len(values_list)
    header = ["type"] + [f"metric_{i}" for i in range(max_cols_needed)]
    
    # 确保文件存在并包含正确的工作表和表头
    if not os.path.exists(file_path):
        wb = Workbook()
        sheet = wb.active
        sheet.title = "ergodic_metric"
        sheet.append(header)
        wb.save(file_path)

    # 加载工作簿
    wb = load_workbook(file_path)
    if "ergodic_metric" in wb.sheetnames:
        sheet = wb["ergodic_metric"]
    else:
        sheet = wb.active
        sheet.title = "ergodic_metric"
        sheet.append(header)

    # 写入新行：decay_type + 所有原始值
    new_row = [str(decay_type)] + [float(v) for v in values_list]
    sheet.append(new_row)

    wb.save(file_path)
    print(f"完整数据已成功追加到 {file_path}（共 {len(values_list)} 个值）")

save_ergodic_metrics_to_excel(global_metric, "baseline3")







